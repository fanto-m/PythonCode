"""
Specifications Models for Manufacturing BOM (Bill of Materials)
models/specification_models.py
"""

from PySide6.QtCore import QObject, Signal, Slot, QAbstractListModel, Qt, QModelIndex
from typing import Optional, List


class SpecificationItemsModel(QAbstractListModel):
    """Модель для позиций спецификации (материалов)"""

    ArticleRole = Qt.ItemDataRole.UserRole + 1
    NameRole = Qt.ItemDataRole.UserRole + 2
    QuantityRole = Qt.ItemDataRole.UserRole + 3
    UnitRole = Qt.ItemDataRole.UserRole + 4
    PriceRole = Qt.ItemDataRole.UserRole + 5
    NotesRole = Qt.ItemDataRole.UserRole + 6

    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db = db_manager
        self.items = []  # List of dicts: {article, name, quantity, unit, price, notes}
        self.current_spec_id = None

    def roleNames(self):
        return {
            self.ArticleRole: b"article",
            self.NameRole: b"name",
            self.QuantityRole: b"quantity",
            self.UnitRole: b"unit",
            self.PriceRole: b"price",
            self.NotesRole: b"notes"
        }

    def rowCount(self, parent=QModelIndex()):
        return len(self.items)

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid() or index.row() >= len(self.items):
            return None

        item = self.items[index.row()]

        if role == self.ArticleRole:
            return item['article']
        elif role == self.NameRole:
            return item['name']
        elif role == self.QuantityRole:
            return item['quantity']
        elif role == self.UnitRole:
            return item['unit']
        elif role == self.PriceRole:
            return item['price']
        elif role == self.NotesRole:
            return item.get('notes', '')

        return None

    def setData(self, index, value, role=Qt.ItemDataRole.EditRole):
        if not index.isValid() or index.row() >= len(self.items):
            return False

        item = self.items[index.row()]

        if role == self.QuantityRole:
            try:
                item['quantity'] = float(value)
                self.dataChanged.emit(index, index, [role])
                return True
            except (ValueError, TypeError):
                return False
        elif role == self.NotesRole:
            item['notes'] = str(value)
            self.dataChanged.emit(index, index, [role])
            return True

        return False

    @Slot(str, str, float, str, float)
    def addItem(self, article, name, quantity, unit, price):
        """Добавляет материал в список"""
        # Check if item already exists
        for existing_item in self.items:
            if existing_item['article'] == article:
                print(f"DEBUG: Item {article} already in specification")
                return

        self.beginInsertRows(QModelIndex(), len(self.items), len(self.items))
        self.items.append({
            'article': article,
            'name': name,
            'quantity': quantity,
            'unit': unit,
            'price': price,
            'notes': ''
        })
        self.endInsertRows()
        print(f"DEBUG: Added item {article} to specification")

    @Slot(int)
    def removeItem(self, index):
        """Удаляет материал из списка"""
        if 0 <= index < len(self.items):
            self.beginRemoveRows(QModelIndex(), index, index)
            removed = self.items.pop(index)
            self.endRemoveRows()
            print(f"DEBUG: Removed item {removed['article']}")

    @Slot()
    def clear(self):
        """Очищает все позиции"""
        self.beginResetModel()
        self.items = []
        self.current_spec_id = None
        self.endResetModel()

    @Slot(int)
    def loadForSpecification(self, spec_id):
        """Загружает позиции для спецификации из БД"""
        self.beginResetModel()
        self.current_spec_id = spec_id
        self.items = []

        db_items = self.db.load_specification_items(spec_id)
        for row in db_items:
            # row: (id, spec_id, article, quantity, notes, name, unit, price, stock)
            self.items.append({
                'id': row[0],
                'article': row[2],
                'name': row[5],
                'quantity': row[3],
                'unit': row[6],
                'price': row[7],
                'notes': row[4] or ''
            })

        self.endResetModel()
        print(f"DEBUG: Loaded {len(self.items)} items for specification {spec_id}")

    @Slot(result=float)
    def getTotalMaterialsCost(self):
        """Возвращает общую стоимость материалов"""
        total = sum(item['quantity'] * item['price'] for item in self.items)
        return total

    def getItems(self):
        """Возвращает список всех позиций для сохранения"""
        return self.items


class SpecificationsModel(QObject):
    """Модель для управления спецификациями"""

    errorOccurred = Signal(str)
    specificationsLoaded = Signal()

    def __init__(self, db_manager, items_model, parent=None):
        super().__init__(parent)
        self.db = db_manager
        self.specification_items_model = items_model  # Use the shared instance

    @Slot(int, str, str, str, float, float, result=int)
    def saveSpecification(self, spec_id, name, description, status, labor_cost, overhead_percentage):
        """
        Сохраняет спецификацию с позициями
        Возвращает ID спецификации или -1 при ошибке
        """
        try:
            if not name or not name.strip():
                self.errorOccurred.emit("Название спецификации не может быть пустым")
                return -1

            items = self.specification_items_model.getItems()

            if not items:
                self.errorOccurred.emit("Спецификация должна содержать хотя бы один материал")
                return -1

            # Convert items to format expected by database
            db_items = [
                {
                    'article': item['article'],
                    'quantity': item['quantity'],
                    'notes': item.get('notes', '')
                }
                for item in items
            ]

            # Save to database (transactionally)
            result_id = self.db.save_specification_with_items(
                spec_id if spec_id > 0 else None,
                name,
                description,
                status,
                labor_cost,
                overhead_percentage,
                db_items
            )

            if result_id:
                print(f"DEBUG: Specification saved with ID: {result_id}")
                return result_id
            else:
                self.errorOccurred.emit("Ошибка при сохранении спецификации")
                return -1

        except Exception as e:
            error_msg = f"Ошибка: {str(e)}"
            print(f"DEBUG: {error_msg}")
            self.errorOccurred.emit(error_msg)
            return -1

    @Slot(int)
    def loadSpecification(self, spec_id):
        """Загружает спецификацию для редактирования"""
        try:
            spec_data = self.db.get_specification(spec_id)
            if spec_data:
                # Load items for this specification
                self.specification_items_model.loadForSpecification(spec_id)
                print(f"DEBUG: Loaded specification {spec_id}")
            else:
                self.errorOccurred.emit(f"Спецификация {spec_id} не найдена")
        except Exception as e:
            self.errorOccurred.emit(f"Ошибка загрузки: {str(e)}")

    @Slot(int, result=bool)
    def deleteSpecification(self, spec_id):
        """Удаляет спецификацию"""
        try:
            success = self.db.delete_specification(spec_id)
            if not success:
                self.errorOccurred.emit("Не удалось удалить спецификацию")
            return success
        except Exception as e:
            self.errorOccurred.emit(f"Ошибка удаления: {str(e)}")
            return False

    @Slot(result="QVariantList")
    def loadAllSpecifications(self):
        """Загружает все спецификации для отображения в списке"""
        try:
            specs = self.db.load_specifications()
            result = []
            for spec in specs:
                result.append({
                    'id': spec[0],
                    'name': spec[1],
                    'description': spec[2] or '',
                    'created_date': spec[3],
                    'modified_date': spec[4],
                    'status': spec[5],
                    'labor_cost': spec[6],
                    'overhead_percentage': spec[7],
                    'final_price': spec[8]
                })
            print(f"DEBUG: Loaded {len(result)} specifications")
            return result
        except Exception as e:
            print(f"DEBUG: Error loading all specifications: {str(e)}")
            self.errorOccurred.emit(f"Ошибка загрузки: {str(e)}")
            return []

    @Slot(int, result="QVariantList")
    def loadSpecificationItems(self, spec_id):
        """Загружает позиции спецификации для отображения"""
        try:
            items = self.db.load_specification_items(spec_id)
            result = []
            for item in items:
                # item: (id, spec_id, article, quantity, notes, name, unit, price, stock)
                result.append({
                    'article': item[2],
                    'name': item[5],
                    'quantity': item[3],
                    'unit': item[6],
                    'price': item[7],
                    'total': item[3] * item[7]
                })
            return result
        except Exception as e:
            print(f"DEBUG: Error loading spec items: {str(e)}")
            return []

    @Slot(int)
    def exportToExcel(self, spec_id):
        """Экспорт спецификации в Excel"""
        try:
            # Get specification data
            spec_data = self.db.get_specification(spec_id)
            if not spec_data:
                self.errorOccurred.emit("Спецификация не найдена")
                return

            items = self.db.load_specification_items(spec_id)
            if not items:
                self.errorOccurred.emit("В спецификации нет позиций")
                return

            # Import openpyxl
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            except ImportError:
                self.errorOccurred.emit("Библиотека openpyxl не установлена. Установите: pip install openpyxl")
                return

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Спецификация"

            # Styles
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Title
            ws['A1'] = f"СПЕЦИФИКАЦИЯ: {spec_data[1]}"
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:G1')

            # Info
            ws['A3'] = f"Описание: {spec_data[2] or ''}"
            ws['A4'] = f"Статус: {spec_data[5]}"
            ws['A5'] = f"Дата создания: {spec_data[3]}"

            # Headers
            headers = ['№', 'Артикул', 'Название', 'Кол-во', 'Ед.', 'Цена', 'Сумма']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=7, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # Data rows
            materials_total = 0
            for idx, item in enumerate(items, 1):
                # item: (id, spec_id, article, quantity, notes, name, unit, price, stock)
                row = idx + 7
                quantity = item[3]
                price = item[7]
                total = quantity * price
                materials_total += total

                ws.cell(row=row, column=1, value=idx).border = border
                ws.cell(row=row, column=2, value=item[2]).border = border  # article
                ws.cell(row=row, column=3, value=item[5]).border = border  # name
                ws.cell(row=row, column=4, value=quantity).border = border
                ws.cell(row=row, column=5, value=item[6]).border = border  # unit
                ws.cell(row=row, column=6, value=price).border = border
                ws.cell(row=row, column=7, value=total).border = border

            # Totals
            total_row = len(items) + 9
            ws.cell(row=total_row, column=6, value="Материалы:").font = Font(bold=True)
            ws.cell(row=total_row, column=7, value=materials_total).font = Font(bold=True)

            ws.cell(row=total_row + 1, column=6, value="Работа:").font = Font(bold=True)
            ws.cell(row=total_row + 1, column=7, value=spec_data[6]).font = Font(bold=True)

            overhead = materials_total * (spec_data[7] / 100)
            ws.cell(row=total_row + 2, column=6, value=f"Накладные ({spec_data[7]}%):").font = Font(bold=True)
            ws.cell(row=total_row + 2, column=7, value=overhead).font = Font(bold=True)

            total = materials_total + spec_data[6] + overhead
            ws.cell(row=total_row + 3, column=6, value="ИТОГО:").font = Font(bold=True, size=14)
            ws.cell(row=total_row + 3, column=7, value=total).font = Font(bold=True, size=14)

            # Column widths
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 8
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12

            # Save file
            filename = f"specification_{spec_id}_{spec_data[1].replace(' ', '_')}.xlsx"
            wb.save(filename)

            print(f"DEBUG: Excel file saved: {filename}")
            self.errorOccurred.emit(f"Файл сохранён: {filename}")

        except Exception as e:
            self.errorOccurred.emit(f"Ошибка экспорта в Excel: {str(e)}")

    @Slot(int)
    def exportToPDF(self, spec_id):
        """Экспорт спецификации в PDF"""
        try:
            # Get specification data
            spec_data = self.db.get_specification(spec_id)
            if not spec_data:
                self.errorOccurred.emit("Спецификация не найдена")
                return

            items = self.db.load_specification_items(spec_id)
            if not items:
                self.errorOccurred.emit("В спецификации нет позиций")
                return

            # Import reportlab
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.lib.units import mm
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.pdfbase import pdfmetrics
                from reportlab.pdfbase.ttfonts import TTFont
            except ImportError:
                self.errorOccurred.emit("Библиотека reportlab не установлена. Установите: pip install reportlab")
                return

            # Create PDF
            filename = f"specification_{spec_id}_{spec_data[1].replace(' ', '_')}.pdf"
            doc = SimpleDocTemplate(filename, pagesize=A4)
            story = []

            styles = getSampleStyleSheet()

            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#366092'),
                spaceAfter=30,
                alignment=1  # Center
            )
            story.append(Paragraph(f"СПЕЦИФИКАЦИЯ: {spec_data[1]}", title_style))

            # Info
            info_style = styles['Normal']
            story.append(Paragraph(f"<b>Описание:</b> {spec_data[2] or 'Не указано'}", info_style))
            story.append(Paragraph(f"<b>Статус:</b> {spec_data[5]}", info_style))
            story.append(Paragraph(f"<b>Дата создания:</b> {spec_data[3]}", info_style))
            story.append(Spacer(1, 20))

            # Table data
            table_data = [['№', 'Артикул', 'Название', 'Кол-во', 'Ед.', 'Цена', 'Сумма']]

            materials_total = 0
            for idx, item in enumerate(items, 1):
                quantity = item[3]
                price = item[7]
                total = quantity * price
                materials_total += total

                table_data.append([
                    str(idx),
                    item[2],  # article
                    item[5],  # name
                    f"{quantity:.2f}",
                    item[6],  # unit
                    f"{price:.2f}",
                    f"{total:.2f}"
                ])

            # Create table
            table = Table(table_data, colWidths=[20, 60, 180, 40, 30, 50, 50])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(table)
            story.append(Spacer(1, 20))

            # Totals
            overhead = materials_total * (spec_data[7] / 100)
            total = materials_total + spec_data[6] + overhead

            totals_data = [
                ['Материалы:', f"{materials_total:.2f} ₽"],
                ['Работа:', f"{spec_data[6]:.2f} ₽"],
                [f'Накладные ({spec_data[7]}%):', f"{overhead:.2f} ₽"],
                ['ИТОГО:', f"{total:.2f} ₽"]
            ]

            totals_table = Table(totals_data, colWidths=[300, 100])
            totals_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 12),
                ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),
            ]))
            story.append(totals_table)

            # Build PDF
            doc.build(story)

            print(f"DEBUG: PDF file saved: {filename}")
            self.errorOccurred.emit(f"Файл сохранён: {filename}")

        except Exception as e:
            self.errorOccurred.emit(f"Ошибка экспорта в PDF: {str(e)}")