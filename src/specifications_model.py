"""
Модель управления спецификациями (BOM - Bill of Materials)
Работает с SpecificationItemsTableModel для отображения позиций
С Repository Pattern и экспортом в PDF/Excel
"""

from PySide6.QtCore import QObject, Signal, Slot
from loguru import logger

from repositories.specifications_repository import SpecificationsRepository
from models.dto import Specification, SpecificationItem


class SpecificationsModel(QObject):
    """
    Модель для управления спецификациями (BOM).

    Координирует работу со спецификациями и их позициями.
    Поддерживает сохранение, загрузку, удаление и экспорт.
    Работает с SpecificationItemsTableModel для управления позициями.
    """

    # Сигналы
    errorOccurred = Signal(str)
    specificationsLoaded = Signal()

    def __init__(
        self,
        specifications_repository: SpecificationsRepository,
        items_table_model,  # SpecificationItemsTableModel
        parent=None
    ):
        """
        Инициализирует модель спецификаций.

        Args:
            specifications_repository: Репозиторий для работы со спецификациями.
            items_table_model: Табличная модель для позиций (SpecificationItemsTableModel).
            parent: Родительский объект Qt (опционально).
        """
        super().__init__(parent)

        self.repository = specifications_repository
        self.specification_items_model = items_table_model

        logger.debug("SpecificationsModel initialized")

    # ==================== CRUD Operations ====================

    @Slot(int, str, str, str, float, float, result=int)
    def saveSpecification(
        self,
        spec_id: int,
        name: str,
        description: str,
        status: str,
        labor_cost: float,
        overhead_percentage: float
    ) -> int:
        """
        Сохраняет спецификацию с позициями.

        Args:
            spec_id: ID спецификации (0 для новой).
            name: Название спецификации.
            description: Описание.
            status: Статус.
            labor_cost: Стоимость работ.
            overhead_percentage: Процент накладных расходов.

        Returns:
            int: ID сохраненной спецификации или -1 при ошибке.
        """
        try:
            # Проверка имени
            if not name or not name.strip():
                self.errorOccurred.emit("Название спецификации не может быть пустым")
                logger.warning("⚠️ Empty specification name")
                return -1

            logger.info(
                f"Saving specification: id={spec_id}, "
                f"name='{name}', status='{status}'"
            )

            # Получаем позиции из табличной модели
            items_data = self.specification_items_model.getAllItems()

            if not items_data:
                self.errorOccurred.emit("Спецификация должна содержать хотя бы один материал")
                logger.warning("⚠️ No items in specification")
                return -1

            logger.debug(f"Got {len(items_data)} items from table model")

            # Вычисляем стоимость материалов
            materials_cost = 0.0
            for item_data in items_data:
                quantity = float(item_data.get('quantity', 0))
                price = float(item_data.get('price', 0))
                materials_cost += quantity * price

            logger.debug(f"Materials cost: {materials_cost:.2f}")

            # Вычисляем накладные и итоговую стоимость
            overhead_cost = materials_cost * (overhead_percentage / 100.0)
            final_price = materials_cost + labor_cost + overhead_cost

            logger.debug(
                f"Costs: materials={materials_cost:.2f}, "
                f"labor={labor_cost:.2f}, overhead={overhead_cost:.2f}, "
                f"final={final_price:.2f}"
            )

            # Создаем DTO для спецификации
            spec = Specification(
                id=spec_id if spec_id > 0 else None,
                name=name,
                description=description,
                status=status,
                labor_cost=labor_cost,
                overhead_percentage=overhead_percentage,
                final_price=final_price,  # Добавляем вычисленную стоимость
                created_date=None,
                modified_date=None
            )

            # Сохраняем спецификацию
            if spec_id == 0 or spec_id <= 0:
                # Создаем новую
                saved_spec_id = self.repository.add(spec)
                logger.info(f"Created new specification with ID: {saved_spec_id}")
            else:
                # Обновляем существующую
                self.repository.update(spec_id, spec)
                saved_spec_id = spec_id
                logger.info(f"Updated specification ID: {saved_spec_id}")

            # Удаляем старые позиции
            self.repository.delete_specification_items(saved_spec_id)
            logger.debug(f"Deleted old items for specification {saved_spec_id}")

            # Сохраняем новые позиции
            for item_data in items_data:
                spec_item = SpecificationItem(
                    id=None,
                    specification_id=saved_spec_id,
                    article=item_data['article'],
                    quantity=float(item_data['quantity']),
                    notes=item_data.get('notes', '')
                )
                self.repository.add_specification_item(spec_item)

            logger.success(
                f"✅ Specification saved: ID={saved_spec_id}, "
                f"items={len(items_data)}"
            )

            return saved_spec_id

        except Exception as e:
            error_msg = f"Ошибка сохранения спецификации: {str(e)}"
            logger.exception(f"❌ {error_msg}")
            self.errorOccurred.emit(error_msg)
            return -1

    @Slot(int, result=bool)
    def deleteSpecification(self, spec_id: int) -> bool:
        """
        Удаляет спецификацию.

        Args:
            spec_id: ID спецификации.

        Returns:
            bool: True если удаление успешно.
        """
        try:
            logger.info(f"Deleting specification {spec_id}")

            self.repository.delete(spec_id)

            logger.success(f"✅ Specification {spec_id} deleted")
            return True

        except Exception as e:
            error_msg = f"Ошибка удаления спецификации: {str(e)}"
            logger.exception(f"❌ {error_msg}")
            self.errorOccurred.emit(error_msg)
            return False

    # ==================== Loading ====================

    @Slot(result="QVariantList")
    def loadAllSpecifications(self):
        """
        Загружает все спецификации для отображения в списке.

        Returns:
            QVariantList: Список словарей с данными спецификаций.
        """
        try:
            logger.info("Loading all specifications")

            specs = self.repository.get_all()

            result = []
            for spec in specs:
                result.append({
                    'id': spec.id,
                    'name': spec.name,
                    'description': spec.description or '',
                    'status': spec.status or 'черновик',
                    'labor_cost': spec.labor_cost or 0.0,
                    'overhead_percentage': spec.overhead_percentage or 0.0,
                    'final_price': spec.final_price or 0.0,  # Добавлено
                    'created_date': spec.created_date or '',
                    'modified_date': spec.modified_date or ''
                })

            logger.success(f"✅ Loaded {len(result)} specifications")
            self.specificationsLoaded.emit()

            return result

        except Exception as e:
            error_msg = f"Ошибка загрузки спецификаций: {str(e)}"
            logger.exception(f"❌ {error_msg}")
            self.errorOccurred.emit(error_msg)
            return []

    @Slot(int, result="QVariantList")
    def loadSpecificationItems(self, spec_id: int):
        """
        Загружает позиции спецификации для табличной модели.

        Args:
            spec_id: ID спецификации.

        Returns:
            QVariantList: Список словарей с данными позиций.
        """
        try:
            logger.info(f"Loading items for specification {spec_id}")

            items = self.repository.get_specification_items(spec_id)

            # Отладка структуры данных
            if items and len(items) > 0:
                logger.debug(f"First item has {len(items[0])} fields")
                logger.trace(f"First item: {items[0]}")

            result = []
            for item in items:
                # Структура из репозитория (13 полей):
                # 0: id, 1: specification_id, 2: article, 3: quantity, 4: notes,
                # 5: name, 6: unit, 7: price, 8: image_path, 9: category,
                # 10: status, 11: manufacturer, 12: description

                quantity = float(item[3]) if len(item) > 3 else 1.0
                price = float(item[7]) if len(item) > 7 and item[7] is not None else 0.0

                item_dict = {
                    'article': item[2] if len(item) > 2 else '',
                    'name': item[5] if len(item) > 5 else '',
                    'quantity': quantity,
                    'unit': item[6] if len(item) > 6 else 'шт.',
                    'price': price,
                    'total': quantity * price,  # Вычисляем сумму
                    'image_path': item[8] if len(item) > 8 else '',
                    'category': item[9] if len(item) > 9 else '',
                    'status': item[10] if len(item) > 10 else '',
                    'notes': item[4] if len(item) > 4 else ''  # Добавлено поле notes
                }
                result.append(item_dict)

            logger.success(
                f"✅ Loaded {len(result)} items for specification {spec_id}"
            )

            return result

        except Exception as e:
            error_msg = f"Ошибка загрузки позиций: {str(e)}"
            logger.exception(f"❌ {error_msg}")
            self.errorOccurred.emit(error_msg)
            return []

    # ==================== Export ====================

    @Slot(int)
    def exportToExcel(self, spec_id: int):
        """
        Экспортирует спецификацию в Excel.

        Args:
            spec_id: ID спецификации.
        """
        try:
            logger.info(f"Exporting specification {spec_id} to Excel")

            # Получаем данные спецификации (теперь возвращается DTO)
            spec = self.repository.get_by_id(spec_id)
            if not spec:
                self.errorOccurred.emit("Спецификация не найдена")
                logger.warning(f"⚠️ Specification {spec_id} not found")
                return

            # Получаем позиции
            items = self.repository.get_specification_items(spec_id)
            if not items:
                self.errorOccurred.emit("В спецификации нет позиций")
                logger.warning(f"⚠️ No items in specification {spec_id}")
                return

            # Проверяем наличие openpyxl
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            except ImportError:
                error_msg = "Библиотека openpyxl не установлена. Установите: pip install openpyxl"
                self.errorOccurred.emit(error_msg)
                logger.error(f"❌ {error_msg}")
                return

            # Создаем книгу
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Спецификация"

            # Стили
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Заголовок
            ws['A1'] = f"СПЕЦИФИКАЦИЯ: {spec.name}"
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:G1')

            # Описание
            description = (spec.description or "").strip()
            description_lines = description.split("\n") if description else [""]

            ws['A3'] = "Описание:"
            ws['A3'].font = Font(bold=True)

            cur_row = 4
            for line in description_lines:
                ws.cell(row=cur_row, column=1, value=line)
                cur_row += 1

            # Статус и дата
            ws.cell(row=cur_row, column=1, value=f"Статус: {spec.status}")
            cur_row += 1
            ws.cell(row=cur_row, column=1, value=f"Дата создания: {spec.created_date or ''}")
            cur_row += 2

            # Заголовки таблицы (№, Название, Кол-во, Ед., Цена, Сумма, Примечание)
            header_row = cur_row
            headers = ['№', 'Название', 'Кол-во', 'Ед.', 'Цена\nбез НДС', 'Сумма\nбез НДС', 'Примечание']

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border

            # Данные
            # Структура items: (id, specification_id, article, quantity, notes,
            #                   name, unit, price, image_path, category, status, manufacturer, description)
            materials_total = 0.0

            for idx, item in enumerate(items, 1):
                row = header_row + idx

                quantity = float(item[3] or 0)  # quantity
                price = float(item[7] or 0)      # price
                line_total = quantity * price
                materials_total += line_total

                ws.cell(row=row, column=1, value=idx).border = border  # №
                ws.cell(row=row, column=2, value=item[5]).border = border  # name
                ws.cell(row=row, column=3, value=quantity).border = border  # quantity
                ws.cell(row=row, column=4, value=item[6]).border = border  # unit

                price_cell = ws.cell(row=row, column=5, value=price)
                price_cell.number_format = '0.00'
                price_cell.border = border

                total_cell = ws.cell(row=row, column=6, value=line_total)
                total_cell.number_format = '0.00'
                total_cell.border = border

                # Примечание (notes из specification_items)
                note_value = item[10] if len(item) > 10 else ''  # notes
                note_cell = ws.cell(row=row, column=7, value=note_value)
                note_cell.border = border

            # Итоги (метки в столбце D, значения в столбце F)
            total_row = header_row + len(items) + 2

            work_cost = float(spec.labor_cost or 0)
            overhead_pct = float(spec.overhead_percentage or 0)
            overhead = materials_total * (overhead_pct / 100.0)
            grand_total = materials_total + work_cost + overhead
            total_with_nds = grand_total * 1.20

            # Метки в столбце D, значения в столбце F
            ws.cell(row=total_row, column=4, value="Материалы:").font = Font(bold=True)
            mat_val = ws.cell(row=total_row, column=6, value=round(materials_total, 2))
            mat_val.number_format = '0.00'
            mat_val.font = Font(bold=True)

            ws.cell(row=total_row + 1, column=4, value="Работа:").font = Font(bold=True)
            work_val = ws.cell(row=total_row + 1, column=6, value=round(work_cost, 2))
            work_val.number_format = '0.00'
            work_val.font = Font(bold=True)

            ws.cell(row=total_row + 2, column=4, value=f"Накладные ({overhead_pct}%):").font = Font(bold=True)
            overhead_val = ws.cell(row=total_row + 2, column=6, value=round(overhead, 2))
            overhead_val.number_format = '0.00'
            overhead_val.font = Font(bold=True)

            ws.cell(row=total_row + 3, column=4, value="ИТОГО:").font = Font(bold=True, size=12)
            total_val = ws.cell(row=total_row + 3, column=6, value=round(grand_total, 2))
            total_val.number_format = '0.00'
            total_val.font = Font(bold=True, size=12)

            # НДС
            ws.cell(row=total_row + 4, column=4, value="ИТОГО с НДС (20%):").font = Font(bold=True, size=12)
            total_nds_cell = ws.cell(row=total_row + 4, column=6, value=round(total_with_nds, 2))
            total_nds_cell.number_format = '0.00'
            total_nds_cell.font = Font(bold=True, size=12)

            # Ширина столбцов
            ws.column_dimensions['A'].width = 5   # №
            ws.column_dimensions['B'].width = 40  # Название
            ws.column_dimensions['C'].width = 10  # Кол-во
            ws.column_dimensions['D'].width = 15  # Ед. и метки итогов
            ws.column_dimensions['E'].width = 12  # Цена
            ws.column_dimensions['F'].width = 12  # Сумма и значения итогов
            ws.column_dimensions['G'].width = 20  # Примечание

            # Сохранение
            filename = f"specification_{spec_id}_{spec.name.replace(' ', '_')}.xlsx"
            wb.save(filename)

            logger.success(f"✅ Excel file saved: {filename}")
            self.errorOccurred.emit(f"Файл сохранён: {filename}")

        except Exception as e:
            error_msg = f"Ошибка экспорта в Excel: {str(e)}"
            logger.exception(f"❌ {error_msg}")
            self.errorOccurred.emit(error_msg)

    @Slot(int, bool)
    def exportToPDF(self, spec_id: int, landscape: bool = False):
        """
        Экспортирует спецификацию в PDF с выбором ориентации.

        Args:
            spec_id: ID спецификации.
            landscape: Альбомная ориентация (по умолчанию False - портретная).
        """
        try:
            logger.info(f"Exporting specification {spec_id} to PDF (landscape={landscape})")

            # Получаем данные спецификации (теперь возвращается DTO)
            spec = self.repository.get_by_id(spec_id)
            if not spec:
                self.errorOccurred.emit("Спецификация не найдена")
                logger.warning(f"⚠️ Specification {spec_id} not found")
                return

            # Получаем позиции
            items = self.repository.get_specification_items(spec_id)
            if not items:
                self.errorOccurred.emit("В спецификации нет позиций")
                logger.warning(f"⚠️ No items in specification {spec_id}")
                return

            # Проверяем ReportLab и PIL
            try:
                from reportlab.lib.pagesizes import A4, landscape as landscape_pagesize
                from reportlab.lib import colors
                from reportlab.lib.units import mm
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.pdfbase import pdfmetrics
                from reportlab.pdfbase.ttfonts import TTFont
                from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
                from reportlab.pdfgen import canvas
                from PIL import Image as PILImage
                from datetime import datetime
            except ImportError as e:
                if 'PIL' in str(e):
                    error_msg = "Библиотека Pillow не установлена. Установите: pip install Pillow"
                else:
                    error_msg = "Библиотека reportlab не установлена. Установите: pip install reportlab"
                self.errorOccurred.emit(error_msg)
                logger.error(f"❌ {error_msg}")
                return

            # Регистрация шрифтов с поддержкой кириллицы
            font_registered = False

            # Вариант 1: GOST type A
            try:
                pdfmetrics.registerFont(TTFont('GOST type A', 'GOST_A_.ttf'))
                pdfmetrics.registerFont(TTFont('GOST type A Bold', 'GOST type A Bold.ttf'))
                font_name = 'GOST type A'
                font_name_bold = 'GOST type A Bold'
                font_registered = True
                logger.debug("Using GOST type A font")
            except:
                pass

            # Вариант 2: FreeSans (Linux)
            if not font_registered:
                try:
                    pdfmetrics.registerFont(TTFont('FreeSans', '/usr/share/fonts/truetype/freefont/FreeSans.ttf'))
                    pdfmetrics.registerFont(TTFont('FreeSans-Bold', '/usr/share/fonts/truetype/freefont/FreeSansBold.ttf'))
                    font_name = 'FreeSans'
                    font_name_bold = 'FreeSans-Bold'
                    font_registered = True
                    logger.debug("Using FreeSans font")
                except:
                    pass

            # Вариант 3: Arial (Windows)
            if not font_registered:
                try:
                    pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
                    pdfmetrics.registerFont(TTFont('Arial-Bold', 'arialbd.ttf'))
                    font_name = 'Arial'
                    font_name_bold = 'Arial-Bold'
                    font_registered = True
                    logger.debug("Using Arial font")
                except:
                    pass

            if not font_registered:
                error_msg = "Не удалось найти шрифт с поддержкой кириллицы"
                self.errorOccurred.emit(error_msg)
                logger.error(f"❌ {error_msg}")
                return

            # Выбор ориентации страницы
            if landscape:
                pagesize = landscape_pagesize(A4)
                orientation_text = "альбомная"
                logger.debug("Using landscape orientation")
            else:
                pagesize = A4
                orientation_text = "портретная"
                logger.debug("Using portrait orientation")

            # Класс для нумерации страниц с общим количеством
            class NumberedCanvas(canvas.Canvas):
                def __init__(self, *args, **kwargs):
                    canvas.Canvas.__init__(self, *args, **kwargs)
                    self._saved_page_states = []

                def showPage(self):
                    self._saved_page_states.append(dict(self.__dict__))
                    self._startPage()

                def save(self):
                    """Добавляем номера страниц после создания всех страниц"""
                    num_pages = len(self._saved_page_states)
                    for state in self._saved_page_states:
                        self.__dict__.update(state)
                        self.draw_page_number(num_pages)
                        canvas.Canvas.showPage(self)
                    canvas.Canvas.save(self)

                def draw_page_number(self, page_count):
                    """Рисуем номер страницы в формате X/Y и дату"""
                    self.saveState()
                    self.setFont(font_name, 9)
                    self.setFillColor(colors.grey)

                    page_num = self._pageNumber

                    # Номер страницы в формате "1/5"
                    page_text = f"{page_num}/{page_count}"
                    self.drawRightString(pagesize[0] - 30, 30, page_text)

                    # Дата под номером страницы
                    current_date = datetime.now().strftime("%d.%m.%Y")
                    self.drawRightString(pagesize[0] - 30, 15, current_date)

                    self.restoreState()

            # Функция для масштабирования изображений
            def scale_image(image_path, max_width_mm=15, max_height_mm=15):
                """Масштабирует изображение с сохранением пропорций"""
                import os
                if not image_path or not os.path.exists(image_path):
                    return None

                try:
                    with PILImage.open(image_path) as pil_img:
                        orig_width, orig_height = pil_img.size

                    max_width = max_width_mm * mm
                    max_height = max_height_mm * mm

                    aspect_ratio = orig_width / orig_height

                    if aspect_ratio > 1:
                        final_width = max_width
                        final_height = max_width / aspect_ratio

                        if final_height > max_height:
                            final_height = max_height
                            final_width = max_height * aspect_ratio
                    else:
                        final_height = max_height
                        final_width = max_height * aspect_ratio

                        if final_width > max_width:
                            final_width = max_width
                            final_height = max_width / aspect_ratio

                    img = Image(image_path, width=final_width, height=final_height)
                    img.hAlign = 'CENTER'

                    logger.trace(f"Image scaled: {orig_width}x{orig_height} -> {final_width:.1f}x{final_height:.1f} pts")
                    return img

                except Exception as e:
                    logger.warning(f"⚠️ Error processing image {image_path}: {e}")
                    return None

            # Функция для добавления шапки
            def add_header(canvas, doc):
                """Добавляет шапку на каждую страницу"""
                canvas.saveState()

                # Шапка
                canvas.setFont(font_name_bold, 14)
                canvas.setFillColor(colors.HexColor('#366092'))

                # Название спецификации с переносом на следующую строку
                canvas.drawString(30, pagesize[1] - 30, "СПЕЦИФИКАЦИЯ:")
                canvas.drawString(30, pagesize[1] - 50, spec.name)

                # Линия под шапкой
                canvas.setStrokeColor(colors.HexColor('#366092'))
                canvas.setLineWidth(2)
                canvas.line(30, pagesize[1] - 60, pagesize[0] - 30, pagesize[1] - 60)

                canvas.restoreState()

            # Создаем PDF
            filename = f"specification_{spec_id}_{spec.name.replace(' ', '_')}.pdf"
            doc = SimpleDocTemplate(
                filename,
                pagesize=pagesize,
                topMargin=80,  # Увеличиваем верхний отступ для шапки
                bottomMargin=50,  # Увеличиваем нижний отступ для даты и номера
                leftMargin=30,
                rightMargin=30
            )
            story = []

            styles = getSampleStyleSheet()

            # Info section
            info_style = ParagraphStyle(
                'CustomInfo',
                parent=styles['Normal'],
                fontName=font_name,
                fontSize=10,
                spaceAfter=6
            )
            story.append(Paragraph(f"<b>Описание:</b> {spec.description or 'Не указано'}", info_style))
            story.append(Paragraph(f"<b>Статус:</b> {spec.status}", info_style))
            story.append(Paragraph(f"<b>Дата создания:</b> {spec.created_date or ''}", info_style))
            story.append(Spacer(1, 15))

            # Адаптивные размеры таблицы
            if landscape:
                # Альбомная: №, Фото, Название, Производитель, Кол-во, Ед., Цена, Сумма, Примечание
                col_widths = [20, 60, 150, 80, 35, 30, 50, 50, 100]
                img_size = 18
            else:
                # Портретная
                col_widths = [15, 50, 110, 70, 30, 25, 45, 45, 80]
                img_size = 15

            # Создание заголовка таблицы с переносом строк
            header_style = ParagraphStyle(
                'TableHeader',
                parent=styles['Normal'],
                fontName=font_name_bold,
                fontSize=9,
                alignment=TA_CENTER,
                leading=10,
                textColor=colors.whitesmoke
            )

            table_data = [[
                Paragraph('№', header_style),
                Paragraph('Изображение', header_style),
                Paragraph('Наименование', header_style),
                Paragraph('Производитель', header_style),
                Paragraph('Кол-во', header_style),
                Paragraph('Ед.', header_style),
                Paragraph('Цена, руб.<br/>без НДС', header_style),
                Paragraph('Сумма, руб.<br/>без НДС', header_style),
                Paragraph('Примечание', header_style)
            ]]

            # Данные
            # Структура items: (id, specification_id, article, quantity, notes,
            #                   name, unit, price, image_path, category, status, manufacturer, description)
            materials_total = 0.0
            for idx, item in enumerate(items, 1):
                quantity = float(item[3])  # quantity
                price = float(item[7])      # price
                total = quantity * price
                materials_total += total

                # Обработка изображения
                image_path = item[8] if len(item) > 8 else ''  # image_path
                image_cell = scale_image(image_path, max_width_mm=img_size, max_height_mm=img_size)

                if image_cell is None:
                    no_image_style = ParagraphStyle(
                        'NoImage',
                        parent=styles['Normal'],
                        fontName=font_name,
                        fontSize=8,
                        alignment=TA_CENTER,
                        textColor=colors.grey
                    )
                    image_cell = Paragraph('Нет<br/>фото', no_image_style)

                # Стили для ячеек
                name_style = ParagraphStyle(
                    'TableName',
                    parent=styles['Normal'],
                    fontName=font_name,
                    fontSize=9,
                    leading=11,
                    alignment=TA_CENTER,
                    wordWrap='CJK',
                )

                center_style = ParagraphStyle(
                    'TableCenter',
                    parent=styles['Normal'],
                    fontName=font_name,
                    fontSize=9,
                    alignment=TA_CENTER,
                )

                note_style = ParagraphStyle(
                    'TableNote',
                    parent=styles['Normal'],
                    fontName=font_name,
                    fontSize=9,
                    leading=11,
                    alignment=TA_CENTER,
                    wordWrap='CJK',
                )

                # Данные из кортежа
                manufacturer = item[11] if len(item) > 11 else ''  # manufacturer
                note = item[10] if len(item) > 10 else ''            # notes (из specification_items)

                table_data.append([
                    Paragraph(str(idx), center_style),
                    image_cell,
                    Paragraph(item[5].replace('\n', '<br/>'), name_style),  # name
                    Paragraph(manufacturer, name_style),
                    Paragraph(f"{quantity:.2f}", center_style),
                    Paragraph(item[6], center_style),  # unit
                    Paragraph(f"{price:.2f}", center_style),
                    Paragraph(f"{total:.2f}", center_style),
                    Paragraph(note, note_style)
                ])

            # Создаем таблицу с адаптивными колонками
            table = Table(table_data, colWidths=col_widths, repeatRows=1)

            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('VALIGN', (8, 1), (8, -1), 'TOP'),
                ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))

            story.append(table)
            story.append(Spacer(1, 20))

            # Итоги
            overhead = materials_total * (spec.overhead_percentage / 100)
            total = materials_total + spec.labor_cost + overhead
            total_with_nds = total * 1.20

            totals_data = [
                ['Материалы:', f"{materials_total:.2f} руб."],
                ['Работа:', f"{spec.labor_cost:.2f} руб."],
                [f'Накладные ({spec.overhead_percentage}%):', f"{overhead:.2f} руб."],
                ['ИТОГО:', f"{total:.2f} руб."],
                ['ИТОГО с НДС (20%):', f"{total_with_nds:.2f} руб."]
            ]

            # Адаптивная ширина таблицы итогов
            totals_width = 500 if landscape else 400
            totals_table = Table(totals_data, colWidths=[totals_width - 100, 100])
            totals_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTNAME', (0, -2), (-1, -2), font_name_bold),
                ('FONTNAME', (0, -1), (-1, -1), font_name_bold),
                ('FONTSIZE', (0, -2), (-1, -2), 12),
                ('FONTSIZE', (0, -1), (-1, -1), 12),
                ('LINEABOVE', (0, -2), (-1, -2), 2, colors.black),
            ]))
            story.append(totals_table)

            # Build PDF с кастомным canvas для нумерации страниц
            doc.build(
                story,
                onFirstPage=add_header,
                onLaterPages=add_header,
                canvasmaker=NumberedCanvas
            )

            logger.success(f"✅ PDF file saved: {filename} ({orientation_text})")
            self.errorOccurred.emit(f"Файл сохранён: {filename} ({orientation_text})")

        except Exception as e:
            error_msg = f"Ошибка экспорта в PDF: {str(e)}"
            logger.exception(f"❌ {error_msg}")
            self.errorOccurred.emit(error_msg)