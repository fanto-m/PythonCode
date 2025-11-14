"""
Specifications Models for Manufacturing BOM (Bill of Materials)
models/specification_models.py
"""
#specification_models.py
from PySide6.QtCore import QObject, Signal, Slot, QAbstractListModel, Qt, QModelIndex
from typing import Optional, List
from PySide6.QtGui import QColor


class SpecificationItemsModel(QAbstractListModel):
    """Модель для позиций спецификации (материалов)"""

    ArticleRole = Qt.ItemDataRole.UserRole + 1
    NameRole = Qt.ItemDataRole.UserRole + 2
    QuantityRole = Qt.ItemDataRole.UserRole + 3
    UnitRole = Qt.ItemDataRole.UserRole + 4
    PriceRole = Qt.ItemDataRole.UserRole + 5
    NotesRole = Qt.ItemDataRole.UserRole + 6
    ImagePathRole = Qt.ItemDataRole.UserRole + 7 #добавлено
    CategoryRole = Qt.ItemDataRole.UserRole + 8 #добавлено
    StatusRole = Qt.ItemDataRole.UserRole + 9

    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db = db_manager
        self.items = []  # List of dicts: {article, name, quantity, unit, price, notes}
        self.current_spec_id = None
        print(f"DEBUG: SpecificationItemsModel initialized at {id(self)}")

    def roleNames(self):
        return {
            self.ArticleRole: b"article",
            self.NameRole: b"name",
            self.QuantityRole: b"quantity",
            self.UnitRole: b"unit",
            self.PriceRole: b"price",
            self.NotesRole: b"notes",
            self.ImagePathRole: b"image_path",#добавлено
            self. CategoryRole: b"category",
            self.StatusRole: b"status",
        }

    def rowCount(self, parent=QModelIndex()):
        return len(self.items)

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid() or index.row() >= len(self.items):
            return None

        item = self.items[index.row()]

        if role == self.ArticleRole:
            return item['article']
        elif role == self.NameRole:
            return item['name']
        elif role == self.QuantityRole:
            return item['quantity']
        elif role == self.UnitRole:
            return item['unit']
        elif role == self.PriceRole:
            return item['price']
        elif role == self.NotesRole:
            return item.get('notes', '')
        elif role == self.ImagePathRole:    #добавлено
            return item.get('image_path', '')
        elif role == self.CategoryRole:
            return item.get('category', '')
        elif role == self.StatusRole:
            return item.get('status', '')
        return None

    def setData(self, index, value, role=Qt.ItemDataRole.EditRole):
        if not index.isValid() or index.row() >= len(self.items):
            return False

        item = self.items[index.row()]

        if role == self.QuantityRole:
            try:
                item['quantity'] = float(value)
                self.dataChanged.emit(index, index, [role])
                return True
            except (ValueError, TypeError):
                return False
        elif role == self.NotesRole:
            item['notes'] = str(value)
            self.dataChanged.emit(index, index, [role])
            return True

        return False

    @Slot(str, str, float, str, float, str, str, str)
    def addItem(self, article, name, quantity, unit, price, image_path, category, status):
        """Добавляет материал в список. Если артикул уже есть — увеличивает количество."""

        print("=" * 80)
        print("=== DEBUG: addItem called ===")
        print(f"Article: {article} (type: {type(article)})")
        print(f"Name: {name} (type: {type(name)})")
        print(f"Quantity: {quantity} (type: {type(quantity)})")
        print(f"Unit: {unit} (type: {type(unit)})")
        print(f"Price: {price} (type: {type(price)})")
        print(f"Image: {image_path} (type: {type(image_path)})")
        print(f"Category: {category} (type: {type(category)})")
        print(f"Status: {status} (type: {type(status)})")
        print(f"Current items count BEFORE: {len(self.items)}")
        print("-" * 80)

        # Нормализуем артикул для сравнения
        article_normalized = str(article).strip()
        print(f"Normalized article for comparison: '{article_normalized}'")

        # ✅ ПРОВЕРЯЕМ, ЕСТЬ ЛИ УЖЕ ТОВАР С ТАКИМ АРТИКУЛОМ
        for i, existing_item in enumerate(self.items):
            existing_article = str(existing_item.get('article', '')).strip()

            print(f"Comparing with item [{i}]:")
            print(f"  Existing article: '{existing_article}'")
            print(f"  New article:      '{article_normalized}'")
            print(f"  Are equal? {existing_article == article_normalized}")

            if existing_article == article_normalized:
                # ✅ НАШЛИ! Увеличиваем количество
                old_quantity = existing_item['quantity']
                existing_item['quantity'] += quantity

                print(f"  ✓ MATCH FOUND! Updating quantity:")
                print(f"    Old quantity: {old_quantity}")
                print(f"    Added: {quantity}")
                print(f"    New quantity: {existing_item['quantity']}")

                # Уведомляем модель об изменении
                index = self.index(i)
                self.dataChanged.emit(index, index, [self.QuantityRole])

                print("=" * 80)
                return  # ⚠️ ВАЖНО: Выходим из функции, не добавляя новый элемент!

        # ✅ НЕ НАШЛИ - ДОБАВЛЯЕМ НОВЫЙ
        print(f"No existing item found with article '{article_normalized}'")
        print("Adding as NEW item...")

        self.beginInsertRows(QModelIndex(), len(self.items), len(self.items))

        new_item = {
            'article': article_normalized,  # Используем нормализованный артикул
            'name': name,
            'quantity': quantity,
            'unit': unit,
            'price': price,
            'image_path': image_path,
            'category': category,
            'status': status,
        }

        self.items.append(new_item)
        self.endInsertRows()

        print(f"✓ New item added successfully")
        print(f"Current items count AFTER: {len(self.items)}")
        print("=== DEBUG: addItem completed ===")
        print("=" * 80)

    @Slot(int)
    def removeItem(self, index):
        """Удаляет материал из списка"""
        if 0 <= index < len(self.items):
            self.beginRemoveRows(QModelIndex(), index, index)
            removed = self.items.pop(index)
            self.endRemoveRows()
            print(f"DEBUG: Removed item {removed['article']}")

    @Slot()
    def clear(self):
        """Очищает все позиции"""
        self.beginResetModel()
        self.items = []
        self.current_spec_id = None
        self.endResetModel()

    @Slot(int)
    def loadForSpecification(self, spec_id):
        """Загружает позиции для спецификации из БД"""
        self.beginResetModel()
        self.current_spec_id = spec_id
        self.items = []

        db_items = self.db.load_specification_items(spec_id)
        for row in db_items:
            # row: (id, spec_id, article, quantity, notes, name, unit, price, image_path)
            self.items.append({
                'id': row[0],
                'article': row[2],
                'name': row[5],
                'quantity': row[3],
                'unit': row[6],
                'price': row[7],
                'notes': row[4] or '',
                'image_path': row[8] or '',#добавлено
                'category': row[9] or '',
            })

        self.endResetModel()
        print(f"DEBUG: Loaded {len(self.items)} items for specification {spec_id}")

    @Slot(result=float)
    def getTotalMaterialsCost(self):
        """Возвращает общую стоимость материалов"""
        total = sum(item['quantity'] * item['price'] for item in self.items)
        return total

    def getItems(self):
        """Возвращает список всех позиций для сохранения"""
        return self.items

    @Slot()
    def debugPrintItems(self):
        """Выводит все элементы модели для отладки"""
        import sys
        print("\n" + "=" * 80, file=sys.stderr, flush=True)
        print(f"DEBUG: Current model state (id={id(self)})", file=sys.stderr, flush=True)
        print(f"Total items: {len(self.items)}", file=sys.stderr, flush=True)
        print("-" * 80, file=sys.stderr, flush=True)

        for i, item in enumerate(self.items):
            print(f"[{i}] Article: '{item.get('article', '')}' | "
                  f"Name: '{item.get('name', '')}' | "
                  f"Qty: {item.get('quantity', 0)} | "
                  f"Unit: '{item.get('unit', '')}' | "
                  f"Price: {item.get('price', 0)}",
                  file=sys.stderr, flush=True)

        print("=" * 80 + "\n", file=sys.stderr, flush=True)


class SpecificationsModel(QObject):
    """Модель для управления спецификациями"""

    errorOccurred = Signal(str)
    specificationsLoaded = Signal()

    def __init__(self, db_manager, items_model, parent=None):
        super().__init__(parent)
        self.db = db_manager
        self.specification_items_model = items_model  # Use the shared instance

    @Slot(int, str, str, str, float, float, result=int)
    def saveSpecification(self, spec_id, name, description, status, labor_cost, overhead_percentage):
        """
        Сохраняет спецификацию с позициями
        Возвращает ID спецификации или -1 при ошибке
        """
        try:
            if not name or not name.strip():
                self.errorOccurred.emit("Название спецификации не может быть пустым")
                return -1

            items = self.specification_items_model.getItems()

            if not items:
                self.errorOccurred.emit("Спецификация должна содержать хотя бы один материал")
                return -1

            # Convert items to format expected by database
            db_items = [
                {
                    'article': item['article'],
                    'quantity': item['quantity'],
                    'notes': item.get('notes', ''),
                    'image_path': item.get('image_path', ''),#добавлено
                    'category': item.get('category', ''),#добавлено
                }
                for item in items
            ]

            # Save to database (transactionally)
            result_id = self.db.save_specification_with_items(
                spec_id if spec_id > 0 else None,
                name,
                description,
                status,
                labor_cost,
                overhead_percentage,
                db_items
            )

            if result_id:
                print(f"DEBUG: Specification saved with ID: {result_id}")
                return result_id
            else:
                self.errorOccurred.emit("Ошибка при сохранении спецификации")
                return -1

        except Exception as e:
            error_msg = f"Ошибка: {str(e)}"
            print(f"DEBUG: {error_msg}")
            self.errorOccurred.emit(error_msg)
            return -1

    @Slot(int)
    def loadSpecification(self, spec_id):
        """Загружает спецификацию для редактирования"""
        try:
            spec_data = self.db.get_specification(spec_id)
            if spec_data:
                # Load items for this specification
                self.specification_items_model.loadForSpecification(spec_id)
                print(f"DEBUG: Loaded specification {spec_id}")
            else:
                self.errorOccurred.emit(f"Спецификация {spec_id} не найдена")
        except Exception as e:
            self.errorOccurred.emit(f"Ошибка загрузки: {str(e)}")

    @Slot(int, result=bool)
    def deleteSpecification(self, spec_id):
        """Удаляет спецификацию"""
        try:
            success = self.db.delete_specification(spec_id)
            if not success:
                self.errorOccurred.emit("Не удалось удалить спецификацию")
            return success
        except Exception as e:
            self.errorOccurred.emit(f"Ошибка удаления: {str(e)}")
            return False

    @Slot(result="QVariantList")
    def loadAllSpecifications(self):
        """Загружает все спецификации для отображения в списке"""
        try:
            specs = self.db.load_specifications()
            result = []
            for spec in specs:
                result.append({
                    'id': spec[0],
                    'name': spec[1],
                    'description': spec[2] or '',
                    'created_date': spec[3],
                    'modified_date': spec[4],
                    'status': spec[5],
                    'labor_cost': spec[6],
                    'overhead_percentage': spec[7],
                    'final_price': spec[8]
                })
            print(f"DEBUG: Loaded {len(result)} specifications")
            return result
        except Exception as e:
            print(f"DEBUG: Error loading all specifications: {str(e)}")
            self.errorOccurred.emit(f"Ошибка загрузки: {str(e)}")
            return []

    @Slot(int, result="QVariantList")
    def loadSpecificationItems(self, spec_id):
        """Загружает позиции спецификации для отображения"""
        try:
            items = self.db.load_specification_items(spec_id)
            result = []

            # ✅ ОТЛАДКА: Проверяем структуру данных
            if items and len(items) > 0:
                print(f"DEBUG: First item has {len(items[0])} fields")
                print(f"DEBUG: First item: {items[0]}")

            for item in items:
                # ✅ Безопасный доступ к полям с проверкой длины
                item_dict = {
                    'article': item[2] if len(item) > 2 else '',
                    'name': item[5] if len(item) > 5 else '',
                    'quantity': float(item[3]) if len(item) > 3 else 1.0,
                    'unit': item[6] if len(item) > 6 else 'шт.',
                    'price': float(item[7]) if len(item) > 7 else 0.0,
                    'total': (float(item[3]) * float(item[7])) if len(item) > 7 else 0.0,
                    'image_path': item[8] if len(item) > 8 else '',
                    'category': item[9] if len(item) > 9 else '',
                    'status': item[10] if len(item) > 10 else ''
                }
                result.append(item_dict)

            print(f"DEBUG: Loaded {len(result)} items for spec {spec_id}")
            return result

        except Exception as e:
            print(f"DEBUG: Error loading spec items: {str(e)}")
            import traceback
            traceback.print_exc()  # ✅ Полный traceback для отладки
            return []

    @Slot(int)
    def exportToExcel(self, spec_id):
        """Экспорт спецификации в Excel"""
        try:
            # Get specification data
            spec_data = self.db.get_specification(spec_id)
            if not spec_data:
                self.errorOccurred.emit("Спецификация не найдена")
                return

            items = self.db.load_specification_items(spec_id)
            if not items:
                self.errorOccurred.emit("В спецификации нет позиций")
                return

            # Import openpyxl
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            except ImportError:
                self.errorOccurred.emit("Библиотека openpyxl не установлена. Установите: pip install openpyxl")
                return

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Спецификация"

            # Styles
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Title
            ws['A1'] = f"СПЕЦИФИКАЦИЯ: {spec_data[1]}"
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:G1')

            # Info
            ws['A3'] = f"Описание: {spec_data[2] or ''}"
            ws['A4'] = f"Статус: {spec_data[5]}"
            ws['A5'] = f"Дата создания: {spec_data[3]}"

            # Headers
            headers = ['№', 'Артикул', 'Название', 'Кол-во', 'Ед.', 'Цена', 'Сумма']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=7, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # Data rows
            materials_total = 0
            for idx, item in enumerate(items, 1):
                # item: (id, spec_id, article, quantity, notes, name, unit, price, stock)
                row = idx + 7
                quantity = item[3]
                price = item[7]
                total = quantity * price
                materials_total += total

                ws.cell(row=row, column=1, value=idx).border = border
                ws.cell(row=row, column=2, value=item[2]).border = border  # article
                ws.cell(row=row, column=3, value=item[5]).border = border  # name
                ws.cell(row=row, column=4, value=quantity).border = border
                ws.cell(row=row, column=5, value=item[6]).border = border  # unit
                ws.cell(row=row, column=6, value=price).border = border
                ws.cell(row=row, column=7, value=total).border = border

            # Totals
            total_row = len(items) + 9
            ws.cell(row=total_row, column=6, value="Материалы:").font = Font(bold=True)
            ws.cell(row=total_row, column=7, value=materials_total).font = Font(bold=True)

            ws.cell(row=total_row + 1, column=6, value="Работа:").font = Font(bold=True)
            ws.cell(row=total_row + 1, column=7, value=spec_data[6]).font = Font(bold=True)

            overhead = materials_total * (spec_data[7] / 100)
            ws.cell(row=total_row + 2, column=6, value=f"Накладные ({spec_data[7]}%):").font = Font(bold=True)
            ws.cell(row=total_row + 2, column=7, value=overhead).font = Font(bold=True)

            total = materials_total + spec_data[6] + overhead
            ws.cell(row=total_row + 3, column=6, value="ИТОГО:").font = Font(bold=True, size=14)
            ws.cell(row=total_row + 3, column=7, value=total).font = Font(bold=True, size=14)

            # Column widths
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 8
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12

            # Save file
            filename = f"specification_{spec_id}_{spec_data[1].replace(' ', '_')}.xlsx"
            wb.save(filename)

            print(f"DEBUG: Excel file saved: {filename}")
            self.errorOccurred.emit(f"Файл сохранён: {filename}")

        except Exception as e:
            self.errorOccurred.emit(f"Ошибка экспорта в Excel: {str(e)}")

    @Slot(int, bool)
    def exportToPDF(self, spec_id, landscape=False):
        """Экспорт спецификации в PDF с выбором ориентации"""
        try:
            # Get specification data
            spec_data = self.db.get_specification(spec_id)
            if not spec_data:
                self.errorOccurred.emit("Спецификация не найдена")
                return

            items = self.db.load_specification_items(spec_id)
            if not items:
                self.errorOccurred.emit("В спецификации нет позиций")
                return

            # Import reportlab
            try:
                from reportlab.lib.pagesizes import A4, landscape as landscape_pagesize
                from reportlab.lib import colors
                from reportlab.lib.units import mm
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.pdfbase import pdfmetrics
                from reportlab.pdfbase.ttfonts import TTFont
                from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
                from reportlab.pdfgen import canvas
                from PIL import Image as PILImage
                from datetime import datetime
            except ImportError as e:
                if 'PIL' in str(e):
                    self.errorOccurred.emit("Библиотека Pillow не установлена. Установите: pip install Pillow")
                else:
                    self.errorOccurred.emit("Библиотека reportlab не установлена. Установите: pip install reportlab")
                return

            # Регистрация шрифтов с поддержкой кириллицы
            font_registered = False

            # Вариант 1: GOST type A
            try:
                pdfmetrics.registerFont(TTFont('GOST type A', 'GOST_A_.ttf'))
                pdfmetrics.registerFont(TTFont('GOST type A Bold', 'GOST type A Bold.ttf'))
                font_name = 'GOST type A'
                font_name_bold = 'GOST type A Bold'
                font_registered = True
                print("DEBUG: Используем шрифт GOST type A")
            except:
                pass

            # Вариант 2: FreeSans (Linux)
            if not font_registered:
                try:
                    pdfmetrics.registerFont(TTFont('FreeSans', '/usr/share/fonts/truetype/freefont/FreeSans.ttf'))
                    pdfmetrics.registerFont(
                        TTFont('FreeSans-Bold', '/usr/share/fonts/truetype/freefont/FreeSansBold.ttf'))
                    font_name = 'FreeSans'
                    font_name_bold = 'FreeSans-Bold'
                    font_registered = True
                    print("DEBUG: Используем шрифт FreeSans")
                except:
                    pass

            # Вариант 3: Arial (Windows)
            if not font_registered:
                try:
                    pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
                    pdfmetrics.registerFont(TTFont('Arial-Bold', 'arialbd.ttf'))
                    font_name = 'Arial'
                    font_name_bold = 'Arial-Bold'
                    font_registered = True
                    print("DEBUG: Используем шрифт Arial")
                except:
                    pass

            if not font_registered:
                self.errorOccurred.emit(
                    "Не удалось найти шрифт с поддержкой кириллицы. "
                    "Установите: pip install reportlab[rlPyCairo]"
                )
                return

            # ВЫБОР ОРИЕНТАЦИИ СТРАНИЦЫ
            if landscape:
                pagesize = landscape_pagesize(A4)
                orientation_text = "альбомная"
                print("DEBUG: Используем альбомную ориентацию")
            else:
                pagesize = A4
                orientation_text = "портретная"
                print("DEBUG: Используем портретную ориентацию")

            # ✅ КЛАСС ДЛЯ НУМЕРАЦИИ СТРАНИЦ С ОБЩИМ КОЛИЧЕСТВОМ
            class NumberedCanvas(canvas.Canvas):
                def __init__(self, *args, **kwargs):
                    canvas.Canvas.__init__(self, *args, **kwargs)
                    self._saved_page_states = []

                def showPage(self):
                    self._saved_page_states.append(dict(self.__dict__))
                    self._startPage()

                def save(self):
                    """Добавляем номера страниц после создания всех страниц"""
                    num_pages = len(self._saved_page_states)
                    for state in self._saved_page_states:
                        self.__dict__.update(state)
                        self.draw_page_number(num_pages)
                        canvas.Canvas.showPage(self)
                    canvas.Canvas.save(self)

                def draw_page_number(self, page_count):
                    """Рисуем номер страницы в формате X/Y и дату"""
                    self.saveState()
                    self.setFont(font_name, 9)
                    self.setFillColor(colors.grey)

                    page_num = self._pageNumber

                    # ✅ Номер страницы в формате "1/5"
                    page_text = f"{page_num}/{page_count}"
                    self.drawRightString(pagesize[0] - 30, 30, page_text)

                    # ✅ Дата под номером страницы
                    current_date = datetime.now().strftime("%d.%m.%Y")
                    self.drawRightString(pagesize[0] - 30, 15, current_date)

                    self.restoreState()

            # ФУНКЦИЯ ДЛЯ МАСШТАБИРОВАНИЯ ИЗОБРАЖЕНИЙ
            def scale_image(image_path, max_width_mm=15, max_height_mm=15):
                """Масштабирует изображение с сохранением пропорций"""
                import os
                if not image_path or not os.path.exists(image_path):
                    return None

                try:
                    with PILImage.open(image_path) as pil_img:
                        orig_width, orig_height = pil_img.size

                    max_width = max_width_mm * mm
                    max_height = max_height_mm * mm

                    aspect_ratio = orig_width / orig_height

                    if aspect_ratio > 1:
                        final_width = max_width
                        final_height = max_width / aspect_ratio

                        if final_height > max_height:
                            final_height = max_height
                            final_width = max_height * aspect_ratio
                    else:
                        final_height = max_height
                        final_width = max_height * aspect_ratio

                        if final_width > max_width:
                            final_width = max_width
                            final_height = max_width / aspect_ratio

                    img = Image(image_path, width=final_width, height=final_height)
                    img.hAlign = 'CENTER'

                    print(
                        f"DEBUG: Изображение масштабировано: {orig_width}x{orig_height} -> {final_width:.1f}x{final_height:.1f} pts")
                    return img

                except Exception as e:
                    print(f"DEBUG: Ошибка обработки изображения {image_path}: {e}")
                    return None

            # ✅ ФУНКЦИЯ ДЛЯ ДОБАВЛЕНИЯ ШАПКИ
            def add_header(canvas, doc):
                """Добавляет шапку на каждую страницу"""
                canvas.saveState()

                # Шапка
                canvas.setFont(font_name_bold, 14)
                canvas.setFillColor(colors.HexColor('#366092'))

                # Название спецификации с переносом на следующую строку
                canvas.drawString(30, pagesize[1] - 30, "СПЕЦИФИКАЦИЯ:")
                canvas.drawString(30, pagesize[1] - 50, spec_data[1])

                # Линия под шапкой
                canvas.setStrokeColor(colors.HexColor('#366092'))
                canvas.setLineWidth(2)
                canvas.line(30, pagesize[1] - 60, pagesize[0] - 30, pagesize[1] - 60)

                canvas.restoreState()

            # Create PDF with custom header/footer
            filename = f"specification_{spec_id}_{spec_data[1].replace(' ', '_')}.pdf"
            doc = SimpleDocTemplate(
                filename,
                pagesize=pagesize,
                topMargin=80,  # Увеличиваем верхний отступ для шапки
                bottomMargin=50,  # Увеличиваем нижний отступ для даты и номера
                leftMargin=30,
                rightMargin=30
            )
            story = []

            styles = getSampleStyleSheet()

            # Info section
            info_style = ParagraphStyle(
                'CustomInfo',
                parent=styles['Normal'],
                fontName=font_name,
                fontSize=10,
                spaceAfter=6
            )
            story.append(Paragraph(f"<b>Описание:</b> {spec_data[2] or 'Не указано'}", info_style))
            story.append(Paragraph(f"<b>Статус:</b> {spec_data[5]}", info_style))
            story.append(Paragraph(f"<b>Дата создания:</b> {spec_data[3]}", info_style))
            story.append(Spacer(1, 15))

            # АДАПТИВНЫЕ РАЗМЕРЫ ТАБЛИЦЫ
            if landscape:
                # Альбомная: №, Фото, Название, Производитель, Кол-во, Ед., Цена, Сумма, Примечание
                col_widths = [20, 60, 150, 80, 35, 30, 50, 50, 100]
                img_size = 18
            else:
                # Портретная
                col_widths = [15, 50, 110, 70, 30, 25, 45, 45, 80]
                img_size = 15

            # СОЗДАНИЕ ЗАГОЛОВКА ТАБЛИЦЫ С ПЕРЕНОСОМ СТРОК
            header_style = ParagraphStyle(
                'TableHeader',
                parent=styles['Normal'],
                fontName=font_name_bold,
                fontSize=9,
                alignment=TA_CENTER,
                leading=10,
                textColor=colors.whitesmoke
            )

            table_data = [[
                Paragraph('№', header_style),
                Paragraph('Изображение', header_style),
                Paragraph('Наименование', header_style),
                Paragraph('Производитель', header_style),
                Paragraph('Кол-во', header_style),
                Paragraph('Ед.', header_style),
                Paragraph('Цена, руб.<br/>без НДС', header_style),
                Paragraph('Сумма, руб.<br/>без НДС', header_style),
                Paragraph('Примечание', header_style)
            ]]

            materials_total = 0
            for idx, item in enumerate(items, 1):
                quantity = item[3]
                price = item[7]
                total = quantity * price
                materials_total += total

                # ПРАВИЛЬНАЯ ОБРАБОТКА ИЗОБРАЖЕНИЯ
                image_path = item[8] if len(item) > 8 else ''
                image_cell = scale_image(image_path, max_width_mm=img_size, max_height_mm=img_size)

                if image_cell is None:
                    no_image_style = ParagraphStyle(
                        'NoImage',
                        parent=styles['Normal'],
                        fontName=font_name,
                        fontSize=8,
                        alignment=TA_CENTER,
                        textColor=colors.grey
                    )
                    image_cell = Paragraph('Нет<br/>фото', no_image_style)

                # Стили для ячеек
                name_style = ParagraphStyle(
                    'TableName',
                    parent=styles['Normal'],
                    fontName=font_name,
                    fontSize=9,
                    leading=11,
                    alignment=TA_CENTER,
                    wordWrap='CJK',
                )


                center_style = ParagraphStyle(
                    'TableCenter',
                    parent=styles['Normal'],
                    fontName=font_name,
                    fontSize=9,
                    alignment=TA_CENTER,
                )

                note_style = ParagraphStyle(
                    'TableNote',
                    parent=styles['Normal'],
                    fontName=font_name,
                    fontSize=9,
                    leading=11,
                    alignment=TA_CENTER,
                    wordWrap='CJK',
                )

                # ✅ ПРАВИЛЬНЫЕ ИНДЕКСЫ для новой структуры
                manufacturer = item[11] if len(item) > 11 else ''  # Производитель
                note = item[10] if len(item) > 10 else ''  # Описание

                # Можно также использовать si.notes если хотите примечание из спецификации:
                # note = item[4]  # si.notes из specification_items

                table_data.append([
                    Paragraph(str(idx), center_style),
                    image_cell,
                    Paragraph(item[5].replace('\n', '<br/>'), name_style),  # i.name
                    Paragraph(manufacturer, name_style),  # i.manufacturer
                    Paragraph(f"{quantity:.2f}", center_style),
                    Paragraph(item[6], center_style),  # i.unit
                    Paragraph(f"{price:.2f}", center_style),
                    Paragraph(f"{total:.2f}", center_style),
                    Paragraph(note, note_style)  # i.description или si.notes
                ])

            # Создаем таблицу с адаптивными колонками
            table = Table(table_data, colWidths=col_widths, repeatRows=1)

            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('VALIGN', (8, 1), (8, -1), 'TOP'),
                ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))

            story.append(table)
            story.append(Spacer(1, 20))

            # Totals
            overhead = materials_total * (spec_data[7] / 100)
            total = materials_total + spec_data[6] + overhead

            totals_data = [
                ['Материалы:', f"{materials_total:.2f} руб."],
                ['Работа:', f"{spec_data[6]:.2f} руб."],
                [f'Накладные ({spec_data[7]}%):', f"{overhead:.2f} руб."],
                ['ИТОГО:', f"{total:.2f} руб."]  # ✅ Исправлено: закрывающая кавычка
            ]

            # Адаптивная ширина таблицы итогов
            totals_width = 500 if landscape else 400
            totals_table = Table(totals_data, colWidths=[totals_width - 100, 100])
            totals_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTNAME', (0, -1), (-1, -1), font_name_bold),
                ('FONTSIZE', (0, -1), (-1, -1), 12),
                ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),
            ]))
            story.append(totals_table)

            # ✅ Build PDF с кастомным canvas для нумерации страниц
            doc.build(
                story,
                onFirstPage=add_header,
                onLaterPages=add_header,
                canvasmaker=NumberedCanvas  # ✅ Используем кастомный canvas
            )

            print(f"DEBUG: PDF file saved: {filename} ({orientation_text})")
            self.errorOccurred.emit(f"Файл сохранён: {filename} ({orientation_text})")

        except Exception as e:
            import traceback
            traceback.print_exc()
            self.errorOccurred.emit(f"Ошибка экспорта в PDF: {str(e)}")